import openpyxl
from pprint import pprint


def dedupe_headers(headers):
    seen = {}
    result = []
    for h in headers:
        if h is None:
            h = "未命名"
        if h in seen:
            seen[h] += 1
            result.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 1
            result.append(h)
    return result


def read_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb["單身資料"]

    headers = dedupe_headers([cell.value for cell in ws[1]])

    raw_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_rows.append(dict(zip(headers, row)))

    # Group lines into vouchers; parent fields only appear on the first line
    vouchers = []
    current = None

    for row in raw_rows:
        if row.get("傳票日期") is not None:
            if current:
                vouchers.append(current)
            current = {
                "voucher_type": row["傳票單別"] or "",
                "voucher_date": row["傳票日期"].strftime("%Y%m%d"),
                "voucher_remark": row["備註"] or "",
                "lines": [],
            }
        if current:
            current["lines"].append({
                "debit_credit": 2 if row["借/貸"] == -1 else row["借/貸"],
                "account_code": row["科目編號"] or "",
                "description": row["摘要"] or "",
                "dept_code": row["部門"] or "",
                "amount": row["金額"] if row["金額"] is not None else "",
                "project_code": row["專案代號"] or "",
                "line_note": row.get("備註_2") or "",
            })

    if current:
        vouchers.append(current)

    return vouchers


if __name__ == "__main__":
    data = read_excel("ACTI10_2.xlsx")
    pprint(data)
